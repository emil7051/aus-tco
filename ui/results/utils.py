"""
Results UI Utilities

UI-specific utilities and constants for TCO results visualization.
This module provides helper functions, data mappings, and formatting tools
that support the visualization of TCO calculation results in the UI layer.
"""

from typing import Dict, List, Any, Optional, Tuple, Union, Callable

from tco_model.models import TCOOutput, ComparisonResult
from tco_model.terminology import (
    UI_COMPONENT_MAPPING, 
    UI_COMPONENT_KEYS, 
    UI_COMPONENT_LABELS,
    get_component_value as get_model_component_value
)
from utils.helpers import format_currency, format_percentage
import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw

# Common cost component mappings
# These represent the standardized cost categories shown across all visualizations
COMPONENT_KEYS = [
    "acquisition",
    "energy", 
    "maintenance",
    "infrastructure",
    "battery_replacement",
    "insurance_registration",  # Combined from insurance + registration
    "taxes_levies",            # Combined from carbon_tax + other_taxes
    "residual_value"
]

# Display labels for each cost component
COMPONENT_LABELS = {
    "acquisition": "Acquisition Costs",
    "energy": "Energy Costs",
    "maintenance": "Maintenance & Repair",
    "infrastructure": "Infrastructure",
    "battery_replacement": "Battery Replacement",
    "insurance_registration": "Insurance & Registration",
    "taxes_levies": "Taxes & Levies",
    "residual_value": "Residual Value"
}

def get_component_value(result: TCOOutput, component: str) -> float:
    """
    Get component NPV value from a result using the standardized access pattern.
    
    Args:
        result: TCO result object containing the cost data
        component: Component key to access
        
    Returns:
        float: The component value in AUD
        
    Example:
        >>> energy_cost = get_component_value(result, "energy")
        >>> insurance_reg = get_component_value(result, "insurance_registration")
    """
    if not result or not result.npv_costs:
        return 0.0
    
    # Use the standardized access function from terminology
    return get_model_component_value(result.npv_costs, component)


def get_annual_component_value(result: TCOOutput, component: str, year: int) -> float:
    """
    Get component value for a specific year using the standardized access pattern.
    
    Args:
        result: TCO result object containing the annual costs data
        component: Component key to access
        year: Year index (0-based)
        
    Returns:
        float: The component value for the specified year in AUD
        
    Example:
        >>> # Get energy costs for year 3
        >>> energy_cost_year_3 = get_annual_component_value(result, "energy", 3)
    """
    if not result or not result.annual_costs or year >= len(result.annual_costs):
        return 0.0
    
    # Use the standardized access function from terminology
    return get_model_component_value(result.annual_costs[year], component)


def get_component_color(component: str) -> str:
    """
    Get the standard color for a component.
    
    Args:
        component: Component key
        
    Returns:
        str: Color hex code for the component
    """
    # Constant for default color
    DEFAULT_COLOR = "#333333"  # Default gray
    
    if component in UI_COMPONENT_MAPPING:
        return UI_COMPONENT_MAPPING[component].get("color", DEFAULT_COLOR)
    return DEFAULT_COLOR


def get_component_display_order(component: str) -> int:
    """
    Get the standard display order for a component.
    
    Args:
        component: Component key
        
    Returns:
        int: Display order value (1-based)
    """
    # Constant for default display order
    DEFAULT_DISPLAY_ORDER = 999  # Default to end
    
    if component in UI_COMPONENT_MAPPING:
        return UI_COMPONENT_MAPPING[component].get("display_order", DEFAULT_DISPLAY_ORDER)
    return DEFAULT_DISPLAY_ORDER


def get_chart_settings(default_settings: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Get chart settings from session state or initialize with defaults.
    
    Maintains chart configuration consistency across different visualizations
    by storing settings in the Streamlit session state.
    
    Args:
        default_settings: Default settings to use if not already in session state
        
    Returns:
        Dict[str, Any]: Dictionary of chart settings
    
    Example:
        >>> settings = get_chart_settings({"chart_height": 600, "show_grid": True})
        >>> chart_height = settings["chart_height"]  # 600
    """
    if default_settings is None:
        default_settings = {
            "show_breakeven_point": True,
            "chart_height": 500,
            "color_scheme": "default",
            "show_grid": True,
            "show_annotations": True,
            "components_to_show": COMPONENT_KEYS.copy()
        }
    
    if "chart_settings" not in st.session_state:
        st.session_state.chart_settings = default_settings
    elif "components_to_show" not in st.session_state.chart_settings:
        # Ensure the components_to_show key exists
        st.session_state.chart_settings["components_to_show"] = COMPONENT_KEYS.copy()
    
    return st.session_state.chart_settings


def apply_chart_theme(fig: Any, height: Optional[int] = None, title: Optional[str] = None) -> Any:
    """
    Apply consistent theme to a plotly figure.
    
    Ensures visual consistency across all charts by applying standardized
    styling, including colors, fonts, margins, and other visual properties.
    
    Args:
        fig: Plotly figure object to style
        height: Chart height in pixels (overrides settings if provided)
        title: Chart title (if provided)
        
    Returns:
        Any: Styled plotly figure
    
    Example:
        >>> fig = px.bar(data, x="year", y="value")
        >>> fig = apply_chart_theme(fig, height=400, title="Annual Costs")
    """
    import streamlit as st
    
    settings = get_chart_settings()
    
    # Base styling
    fig.update_layout(
        height=height or settings["chart_height"],
        font=dict(family="Arial, sans-serif"),
        plot_bgcolor="white",
        margin=dict(l=50, r=50, t=50, b=50),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        )
    )
    
    # Add title if provided
    if title:
        fig.update_layout(title=title)
    
    # Style axes - control grid visibility through axes properties instead
    fig.update_xaxes(showgrid=settings["show_grid"], zeroline=True, zerolinewidth=1, zerolinecolor="lightgray")
    fig.update_yaxes(showgrid=settings["show_grid"], zeroline=True, zerolinewidth=1, zerolinecolor="lightgray")
    
    return fig


def validate_tco_results(results: Dict[str, TCOOutput]) -> bool:
    """
    Validate that TCO results are available and contain necessary data.
    
    Args:
        results: Dictionary containing TCO results for vehicles
        
    Returns:
        bool: Whether the results are valid
    """
    if not results:
        return False
        
    # Check if results contain both vehicles
    if "vehicle_1" not in results or "vehicle_2" not in results:
        return False
        
    # Check if the result objects are valid
    if not results["vehicle_1"] or not results["vehicle_2"]:
        return False
        
    # Check that each result has the required attributes
    for vehicle in ["vehicle_1", "vehicle_2"]:
        result = results[vehicle]
        required_attrs = [
            "total_tco", "lcod", "analysis_period_years", 
            "vehicle_name", "total_distance_km"
        ]
        
        for attr in required_attrs:
            if not hasattr(result, attr):
                return False
    
    return True


def generate_results_export(results, comparison, include_emissions=True, include_charts=True):
    """
    Generate Excel export with all TCO model data
    
    Args:
        results: Dictionary of actual TCO results
        comparison: Actual comparison result object
        include_emissions: Whether to include emissions data in the export
        include_charts: Whether to include chart images in the export
        
    Returns:
        Excel file content as bytes
    """
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create sheets
    summary_sheet = wb.create_sheet("Summary")
    annual_sheet = wb.create_sheet("Annual Costs")
    components_sheet = wb.create_sheet("Cost Components")
    
    # Only include emissions if requested
    if include_emissions:
        emissions_sheet = wb.create_sheet("Emissions")
    
    params_sheet = wb.create_sheet("Parameters")
    
    # Add charts sheet if requested
    if include_charts:
        charts_sheet = wb.create_sheet("Charts")
    
    # Get results
    result1 = results["vehicle_1"]
    result2 = results["vehicle_2"]
    
    # --- Summary Sheet ---
    summary_data = [
        ["TCO Analysis Results", "", ""],
        ["", "", ""],
        ["Metric", result1.vehicle_name, result2.vehicle_name],
        ["Total TCO", result1.total_tco, result2.total_tco],
        ["Cost per km", result1.lcod, result2.lcod],
    ]
    
    # Only include emissions in summary if requested
    if include_emissions:
        summary_data.extend([
            ["Total CO2 (tonnes)", result1.emissions.total_co2_tonnes if hasattr(result1, 'emissions') else "N/A", 
                             result2.emissions.total_co2_tonnes if hasattr(result2, 'emissions') else "N/A"],
            ["CO2 per km (g/km)", result1.emissions.co2_per_km if hasattr(result1, 'emissions') else "N/A", 
                             result2.emissions.co2_per_km if hasattr(result2, 'emissions') else "N/A"],
        ])
    
    summary_data.extend([
        ["", "", ""],
        ["Comparison", "", ""],
        ["Cheaper option", comparison.cheaper_option == 1 and result1.vehicle_name or result2.vehicle_name, ""],
        ["TCO difference", abs(comparison.tco_difference), f"{abs(comparison.tco_percentage):.1f}%"],
        ["", "", ""],
        ["Investment Analysis", "", ""],
    ])
    
    # Add investment analysis if available
    if hasattr(comparison, 'investment_analysis') and comparison.investment_analysis:
        investment = comparison.investment_analysis
        summary_data.extend([
            ["Payback period", investment.has_payback and f"{investment.payback_years:.1f} years" or "No payback", ""],
            ["ROI", investment.roi and f"{investment.roi:.1f}%" or "N/A", ""],
            ["IRR", investment.irr and f"{investment.irr:.1f}%" or "N/A", ""],
        ])
    
    # Add summary data to sheet
    for row in summary_data:
        summary_sheet.append(row)
    
    # --- Annual Costs Sheet ---
    years = list(range(1, max(len(result1.annual_costs), len(result2.annual_costs)) + 1))
    annual_data = {
        "Year": years,
        f"{result1.vehicle_name} Costs": result1.annual_costs.total + [0] * (len(years) - len(result1.annual_costs)),
        f"{result2.vehicle_name} Costs": result2.annual_costs.total + [0] * (len(years) - len(result2.annual_costs)),
    }
    annual_df = pd.DataFrame(annual_data)
    
    # Add to sheet
    for row in dataframe_to_rows(annual_df, index=False, header=True):
        annual_sheet.append(row)
    
    # --- Cost Components Sheet ---
    from tco_model.calculator import TCOCalculator
    calculator = TCOCalculator()
    from tco_model.terminology import UI_COMPONENT_KEYS, UI_COMPONENT_LABELS
    
    # Get component values
    component_data = {"Component": [UI_COMPONENT_LABELS.get(k, k) for k in UI_COMPONENT_KEYS]}
    component_data[result1.vehicle_name] = [calculator.get_component_value(result1, k) for k in UI_COMPONENT_KEYS]
    component_data[result2.vehicle_name] = [calculator.get_component_value(result2, k) for k in UI_COMPONENT_KEYS]
    component_data["Difference"] = [v2 - v1 for v1, v2 in zip(component_data[result1.vehicle_name], component_data[result2.vehicle_name])]
    
    component_df = pd.DataFrame(component_data)
    
    # Add to sheet
    for row in dataframe_to_rows(component_df, index=False, header=True):
        components_sheet.append(row)
    
    # --- Emissions Sheet (if requested) ---
    if include_emissions:
        if hasattr(result1, 'emissions') and hasattr(result2, 'emissions'):
            emissions_data = {
                "Year": years,
                f"{result1.vehicle_name} CO2 (tonnes)": result1.emissions.annual_co2_tonnes + [0] * (len(years) - len(result1.emissions.annual_co2_tonnes)),
                f"{result2.vehicle_name} CO2 (tonnes)": result2.emissions.annual_co2_tonnes + [0] * (len(years) - len(result2.emissions.annual_co2_tonnes)),
            }
            
            # Add cumulative emissions
            emissions_data[f"{result1.vehicle_name} Cumulative CO2"] = np.cumsum(emissions_data[f"{result1.vehicle_name} CO2 (tonnes)"])
            emissions_data[f"{result2.vehicle_name} Cumulative CO2"] = np.cumsum(emissions_data[f"{result2.vehicle_name} CO2 (tonnes)"])
            
            emissions_df = pd.DataFrame(emissions_data)
            
            # Add emissions summary
            emissions_summary = [
                ["", "", ""],
                ["Summary Metrics", result1.vehicle_name, result2.vehicle_name],
                ["Total CO2 (tonnes)", result1.emissions.total_co2_tonnes, result2.emissions.total_co2_tonnes],
                ["CO2 per km (g/km)", result1.emissions.co2_per_km, result2.emissions.co2_per_km],
                ["Energy consumption (kWh)", result1.emissions.energy_consumption_kwh, result2.emissions.energy_consumption_kwh],
                ["Energy per km (kWh/km)", result1.emissions.energy_per_km, result2.emissions.energy_per_km],
                ["Trees equivalent", result1.emissions.trees_equivalent, result2.emissions.trees_equivalent],
                ["Homes equivalent", result1.emissions.homes_equivalent, result2.emissions.homes_equivalent],
                ["Cars equivalent", result1.emissions.cars_equivalent, result2.emissions.cars_equivalent],
            ]
        else:
            # Calculate emissions data using TCO calculator if not available in results
            from tco_model.calculator import TCOCalculator
            calculator = TCOCalculator()
            
            # Generate emissions data based on vehicle types and parameters
            # This ensures we always have emissions data, even for older TCO results
            estimated_emissions1 = calculator.estimate_emissions(result1)
            estimated_emissions2 = calculator.estimate_emissions(result2)
            
            # Create emissions data with calculated values
            emissions_data = {
                "Year": years,
                f"{result1.vehicle_name} CO2 (tonnes)": estimated_emissions1.annual_co2_tonnes + [0] * (len(years) - len(estimated_emissions1.annual_co2_tonnes)),
                f"{result2.vehicle_name} CO2 (tonnes)": estimated_emissions2.annual_co2_tonnes + [0] * (len(years) - len(estimated_emissions2.annual_co2_tonnes)),
            }
            
            # Add cumulative emissions
            emissions_data[f"{result1.vehicle_name} Cumulative CO2"] = np.cumsum(emissions_data[f"{result1.vehicle_name} CO2 (tonnes)"])
            emissions_data[f"{result2.vehicle_name} Cumulative CO2"] = np.cumsum(emissions_data[f"{result2.vehicle_name} CO2 (tonnes)"])
            
            emissions_df = pd.DataFrame(emissions_data)
            
            # Add emissions summary with calculated values
            emissions_summary = [
                ["", "", ""],
                ["Summary Metrics", result1.vehicle_name, result2.vehicle_name],
                ["Total CO2 (tonnes)", estimated_emissions1.total_co2_tonnes, estimated_emissions2.total_co2_tonnes],
                ["CO2 per km (g/km)", estimated_emissions1.co2_per_km, estimated_emissions2.co2_per_km],
                ["Energy consumption (kWh)", estimated_emissions1.energy_consumption_kwh, estimated_emissions2.energy_consumption_kwh],
                ["Energy per km (kWh/km)", estimated_emissions1.energy_per_km, estimated_emissions2.energy_per_km],
                ["Trees equivalent", estimated_emissions1.trees_equivalent, estimated_emissions2.trees_equivalent],
                ["Homes equivalent", estimated_emissions1.homes_equivalent, estimated_emissions2.homes_equivalent],
                ["Cars equivalent", estimated_emissions1.cars_equivalent, estimated_emissions2.cars_equivalent],
            ]
        
        # Add to sheet
        for row in dataframe_to_rows(emissions_df, index=False, header=True):
            emissions_sheet.append(row)
        
        # Add emissions summary
        for i, row in enumerate(emissions_summary):
            for j, value in enumerate(row):
                emissions_sheet.cell(row=len(emissions_df) + 3 + i, column=j+1, value=value)
    
    # --- Charts Sheet (if requested) ---
    if include_charts:
        from ui.results.charts import create_cost_breakdown_chart, create_cumulative_tco_chart
        import plotly.graph_objects as go
        
        # Row counter for charts
        chart_row = 1
        
        # Add cost breakdown chart
        charts_sheet.cell(row=chart_row, column=1, value="Cost Breakdown")
        chart_row += 1
        
        # Create and save cost breakdown chart
        cost_chart = create_cost_breakdown_chart(result1)
        img_bytes = export_chart_as_image(cost_chart, "png")
        
        # Add image to workbook
        img = Image(io.BytesIO(img_bytes))
        img.width = 500
        img.height = 300
        charts_sheet.add_image(img, f'A{chart_row}')
        chart_row += 20  # Leave space for the image
        
        # Add cumulative TCO chart
        charts_sheet.cell(row=chart_row, column=1, value="Cumulative TCO")
        chart_row += 1
        
        # Create and save cumulative TCO chart
        cumul_chart = create_cumulative_tco_chart(result1, result2, comparison)
        img_bytes = export_chart_as_image(cumul_chart, "png")
        
        # Add image to workbook
        img = Image(io.BytesIO(img_bytes))
        img.width = 500
        img.height = 300
        charts_sheet.add_image(img, f'A{chart_row}')
    
    # --- Parameters Sheet ---
    # Extract parameters from scenarios
    scenario1 = result1.scenario
    scenario2 = result2.scenario
    
    if scenario1 and scenario2:
        # Create parameter sections
        param_sections = []
        
        # Vehicle parameters
        if hasattr(scenario1, 'vehicle') and hasattr(scenario2, 'vehicle'):
            vehicle_params = extract_comparable_attributes(scenario1.vehicle, scenario2.vehicle)
            param_sections.append(("Vehicle Parameters", vehicle_params))
        
        # Operational parameters
        if hasattr(scenario1, 'operational') and hasattr(scenario2, 'operational'):
            operational_params = extract_comparable_attributes(scenario1.operational, scenario2.operational)
            param_sections.append(("Operational Parameters", operational_params))
        
        # Economic parameters
        if hasattr(scenario1, 'economic') and hasattr(scenario2, 'economic'):
            economic_params = extract_comparable_attributes(scenario1.economic, scenario2.economic)
            param_sections.append(("Economic Parameters", economic_params))
        
        # If no structured parameters were found, extract top-level attributes
        if not param_sections:
            top_level_params = extract_comparable_attributes(scenario1, scenario2)
            param_sections.append(("Scenario Parameters", top_level_params))
            
        # Add all parameter sections to the sheet
        current_row = 1
        for section_name, params in param_sections:
            # Add section header
            params_sheet.cell(row=current_row, column=1, value=section_name)
            current_row += 1
            
            # Add parameter headers
            params_sheet.cell(row=current_row, column=1, value="Parameter")
            params_sheet.cell(row=current_row, column=2, value=result1.vehicle_name)
            params_sheet.cell(row=current_row, column=3, value=result2.vehicle_name)
            current_row += 1
            
            # Add parameters
            for param_name, (value1, value2) in params.items():
                params_sheet.cell(row=current_row, column=1, value=param_name)
                params_sheet.cell(row=current_row, column=2, value=value1)
                params_sheet.cell(row=current_row, column=3, value=value2)
                current_row += 1
            
            # Add blank row between sections
            current_row += 1
    else:
        params_sheet.cell(row=1, column=1, value="Parameter data not available")
    
    # Apply styling (header fonts, column widths, etc.)
    for sheet in wb.worksheets:
        # Set column widths
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].width = 20
        
        # Style headers
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def export_chart_as_image(chart, img_format="png"):
    """
    Export a plotly chart as an image.
    
    Args:
        chart: Plotly chart figure
        img_format: Image format (png, jpeg, svg, pdf)
        
    Returns:
        Bytes containing the image data
    """
    import plotly.io as pio
    
    # Set the image dimensions
    width = 800
    height = 600
    
    # The following exports may raise exceptions if required dependencies
    # are not installed (like kaleido for PNG export)
    try:
        # Try to export as bytes directly
        img_bytes = pio.to_image(chart, format=img_format, width=width, height=height)
        return img_bytes
    except Exception as e:
        # Fallback to base64 and decode
        try:
            import base64
            img_base64 = pio.to_image(chart, format=img_format, width=width, height=height, engine="kaleido")
            return img_base64
        except Exception as sub_e:
            # If all export methods fail, return a simple error image
            # This avoids breaking the export completely
            return create_error_image(str(e))


def create_error_image(error_message):
    """
    Create a simple error image when chart export fails.
    
    Args:
        error_message: Error message to display
        
    Returns:
        Bytes containing a simple error image
    """
    try:
        # Create a blank image with white background
        width, height = 400, 200
        image = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(image)
        
        # Add error message
        draw.text((20, 20), "Chart Export Error", fill='red')
        draw.text((20, 50), error_message[:50], fill='black')
        
        # Save to bytes
        buffer = io.BytesIO()
        image.save(buffer, format='PNG')
        buffer.seek(0)
        return buffer.getvalue()
    except Exception:
        # Ultimate fallback - return empty bytes if even PIL fails
        return b''


def extract_comparable_attributes(obj1, obj2):
    """
    Extract comparable attributes from two objects for parameter comparison.
    
    Args:
        obj1: First object
        obj2: Second object
        
    Returns:
        Dictionary with attribute names as keys and tuples of (value1, value2) as values
    """
    attributes = {}
    
    # Get all attributes from both objects
    all_attrs = set(dir(obj1)).intersection(set(dir(obj2)))
    
    for attr in all_attrs:
        # Skip private attributes, methods, and callable attributes
        if attr.startswith('_') or attr == 'model_config':
            continue
        
        try:
            # Try to get the attribute values
            value1 = getattr(obj1, attr)
            
            # Skip callable attributes
            if callable(value1):
                continue
                
            value2 = getattr(obj2, attr)
            
            # Only include simple values (skip nested objects)
            if isinstance(value1, (int, float, str, bool)) and isinstance(value2, (int, float, str, bool)):
                attributes[attr] = (value1, value2)
        except AttributeError:
            # Skip attributes that raise AttributeError
            continue
    
    return attributes


def get_column_letter(col_idx):
    """
    Convert a column index to column letter (A, B, C, ..., Z, AA, AB, ...)
    
    Args:
        col_idx: 1-based column index
        
    Returns:
        Column letter(s)
    """
    column_letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    result = ""
    
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = column_letters[remainder] + result
    
    return result


def format_result_for_session_state(result: TCOOutput) -> Dict[str, Any]:
    """
    Format TCO result for storage in session state.
    
    Args:
        result: TCO result object
        
    Returns:
        Dict[str, Any]: Formatted result dictionary
    """
    # This function converts a TCO result to a dictionary
    # that can be stored in session state and is JSON serializable
    
    # Basic properties
    formatted = {
        "vehicle_name": result.vehicle_name,
        "analysis_period_years": result.analysis_period_years,
        "total_distance_km": result.total_distance_km,
        "total_tco": float(result.total_tco),
        "lcod": float(result.lcod)
    }
    
    # Cost components
    for comp in UI_COMPONENT_KEYS:
        formatted[f"{comp}_cost"] = float(get_component_value(result, comp))
    
    # Annual costs if available
    if hasattr(result, "annual_costs"):
        formatted["annual_costs"] = {}
        for year, annual_cost in result.annual_costs.items():
            formatted["annual_costs"][year] = {
                "total": float(annual_cost.total) if hasattr(annual_cost, "total") else 0
            }
            
            # Add component costs if available
            for comp in UI_COMPONENT_KEYS:
                value = get_annual_component_value(result, comp, year)
                formatted["annual_costs"][year][comp] = float(value)
    
    return formatted